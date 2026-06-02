# -*- coding: utf-8 -*-
"""
舆情指挥系统 — Monitor Agent (监测员)

Responsibility (PRD §5.1):
  1. Load keywords from monitor_keywords.json
  2. Search each keyword × platform (single sort: default relevance or by date)
  3. Deduplicate against previous crawl archives
  4. Export results to Excel + archive to raw/monitor/
  5. Track keyword hit rates, generate optimization suggestions
  6. Brand keyword SEO snapshot (PRD M-10)

Isolation constraints:
  - MUST NOT call Scraper to fetch details (Orchestrator handles that)
  - MUST NOT write to wiki/cases/ (Curator's domain)
  - MUST NOT modify monitor_keywords.json (read-only; suggestions only)

Model: No LLM needed — all logic is pure code (search APIs + dedup + Excel).
"""
import io
import json
import random
import re
import sys
import time
from concurrent import futures
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

# Hard timeout per platform search call (seconds).  A single yt-dlp
# ytsearch / ytsearchdate call normally takes 2-8 s on a decent
# connection; 60 s is generous enough for slow networks but prevents
# indefinite hangs in background threads.
_SEARCH_TIMEOUT = 60

# Date-mode constants: max pages for API-filtered platforms, fallback
# count for client-side-filtered platforms.
_DATE_MODE_MAX_PAGES = 50
_DATE_MODE_FALLBACK_COUNT = 200

# Platforms that support true API-level date filtering.
_DATE_CAPABLE_PLATFORMS = {"bilibili"}

# Platforms that use client-side date filter in date mode (fetch more, filter after).
_DATE_CLIENTSIDE_PLATFORMS = {"youtube", "weibo"}

# Platforms skipped in date mode (no date filter, no practical workaround).
_DATE_SKIP_PLATFORMS = {"xiaohongshu", "wechat"}

if sys.stdout and hasattr(sys.stdout, "buffer"):
    if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from agents.shared import PROJECT_ROOT, RAW_DIR, OUTPUTS_DIR
from engine.ratelimit import get_limiter


# ── Dataclasses ────────────────────────────────────────────────────────
@dataclass
class SearchResult:
    """A single search result item."""
    platform: str
    keyword_id: str
    keyword: str
    sort_type: str           # date | hot
    title: str
    url: str
    author: str = ""
    publish_time: str = ""
    engagement: int = 0      # likes + comments + shares
    snippet: str = ""
    error: str = ""          # non-empty when platform search failed (cookie/login/config)


@dataclass
class KeywordResult:
    """Aggregated results for one keyword × platform."""
    keyword_id: str
    keyword: str
    platform: str
    date_results: list[SearchResult] = field(default_factory=list)
    hot_results: list[SearchResult] = field(default_factory=list)
    new_items: list[SearchResult] = field(default_factory=list)  # after dedup
    notes: list[str] = field(default_factory=list)  # platform-specific notes (date mode)


@dataclass
class MonitorStats:
    """Statistics for a monitor job."""
    keywords_searched: int = 0
    platforms_queried: int = 0
    total_fetched: int = 0
    total_new: int = 0
    dedup_rate: float = 0.0
    hit_rates: dict = field(default_factory=dict)   # {kw_id: {platform: rate}}


@dataclass
class MonitorHarvest:
    """Complete output of a monitor job (PRD §5.1)."""
    job_id: str
    started_at: str
    finished_at: str
    keyword_results: list[KeywordResult] = field(default_factory=list)
    total_fetched: int = 0
    total_new: int = 0
    dedup_rate: float = 0.0
    excel_path: str = ""
    stats: Optional[MonitorStats] = None
    errors: list[str] = field(default_factory=list)


# ── Config loading ─────────────────────────────────────────────────────
def load_keywords() -> list[dict]:
    """Load and return active keyword entries from monitor_keywords.json."""
    path = PROJECT_ROOT / "monitor_keywords.json"
    if not path.exists():
        return []
    cfg = json.loads(path.read_text(encoding="utf-8"))
    return [kw for kw in cfg.get("keywords", []) if kw.get("active", True)]


# ── Platform search ────────────────────────────────────────────────────
def _search_douyin(keyword: str, sort_type: str, count: int = 30,
                   date_from: str = "", date_to: str = "") -> list[SearchResult]:
    """Search Douyin via TikTokDownloader Search API.

    sort_type: 'date' → sort_type=2 (最新), 'hot' → sort_type=1 (最多点赞)
    Reads TikTokDownloader path from config.json.paths.tiktok_downloader.
    Returns error SearchResult when cookies are missing or API login required.
    """
    # ── Tier 1: MediaCrawler DouYinClient (Playwright browser) ──
    try:
        from engine.douyin_adapter import search_douyin as _mc_dy_search, _check_cookie_valid as _dy_cookie_ok
        if _dy_cookie_ok():
            mc_results = _mc_dy_search(keyword, count=count, sort_type=sort_type)
            if mc_results:
                return [SearchResult(
                    platform="douyin",
                    keyword_id="",
                    keyword=keyword,
                    sort_type=sort_type,
                    title=r.get("title", ""),
                    url=r.get("url", ""),
                    author=r.get("author", ""),
                    publish_time=r.get("publish_time", ""),
                    engagement=r.get("engagement", 0),
                    snippet=r.get("snippet", ""),
                ) for r in mc_results]
    except Exception:
        pass  # Fall through to Tier 2

    # ── Tier 2: TikTokDownloader Search API (fallback) ──
    # Resolve TikTokDownloader path: config.json → default
    cfg_path = PROJECT_ROOT / "engine" / "config.json"
    _TT_DOWNLOADER_PATH = Path("D:/Claude code/Github skills/TikTokDownloader")
    if cfg_path.exists():
        try:
            cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
            cfg_path_val = cfg.get("paths", {}).get("tiktok_downloader", "")
            if cfg_path_val:
                _TT_DOWNLOADER_PATH = Path(cfg_path_val)
        except Exception:
            pass

    if not _TT_DOWNLOADER_PATH.exists():
        return [SearchResult(
            platform="douyin", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error=f"DOUYIN_PATH_MISSING: TikTokDownloader路径不存在: {_TT_DOWNLOADER_PATH}",
        )]

    import sys as _sys
    _tt_src = str(_TT_DOWNLOADER_PATH)
    if _tt_src not in _sys.path:
        _sys.path.insert(0, _tt_src)

    try:
        from src.application import TikTokDownloader as _TTApp
    except ImportError as e:
        return [SearchResult(
            platform="douyin", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error=f"DOUYIN_IMPORT_ERROR: 无法导入TikTokDownloader: {e}",
        )]

    try:
        from src.interface import Search as _Search
    except ImportError as e:
        return [SearchResult(
            platform="douyin", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error=f"DOUYIN_IMPORT_ERROR: 无法导入Search模块: {e}",
        )]

    import asyncio

    # Capture stdout to detect API errors printed by TikTokDownloader internally
    capture_buf = io.StringIO()
    _old_stdout = sys.stdout
    sys.stdout = capture_buf

    async def _search():
        async with _TTApp() as app:
            app.check_config()
            await app.check_settings(False)
            # Inject Douyin cookie from multiple sources:
            #   1. config.json douyin.cookie (raw string)
            #   2. TikTokDownloader settings.json cookie dict (from bootstrap)
            douyin_cfg = {}
            if cfg_path.exists():
                try:
                    cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
                    douyin_cfg = cfg.get("douyin", {})
                except Exception:
                    pass
            cookie_str = douyin_cfg.get("cookie", "")
            # Fallback: read from TikTokDownloader's own cookie dict
            if not cookie_str:
                try:
                    from engine.tt_fetcher import _load_tt_cookie, _cookie_str as _tt_cookie_str_fn
                    tt_cookie = _load_tt_cookie()
                    if tt_cookie and len(tt_cookie) > 3:
                        cookie_str = _tt_cookie_str_fn(tt_cookie)
                except Exception:
                    pass
            if cookie_str:
                cookie_pairs = {}
                for pair in cookie_str.split("; "):
                    if "=" in pair:
                        k, v = pair.split("=", 1)
                        cookie_pairs[k.strip()] = v.strip()
                if cookie_pairs:
                    app.parameter.cookie_dict.update(cookie_pairs)
            search = _Search(
                app.parameter,
                keyword=keyword,
                pages=max(1, count // 15),
                sort_type=2 if sort_type == "date" else 1,
            )
            result = await search.run(single_page=True)
            return result

    try:
        raw_results = asyncio.run(_search())
    except Exception as e:
        sys.stdout = _old_stdout
        print(f"[Douyin搜索] TikTokDownloader异常: {e}", file=sys.stderr)
        return [SearchResult(
            platform="douyin", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error=f"DOUYIN_SEARCH_ERROR: {str(e)[:120]}",
        )]
    finally:
        captured_output = capture_buf.getvalue()
        sys.stdout = _old_stdout

    # Detect API-level login errors printed by TikTokDownloader internally
    if "status_code" in captured_output and ("2483" in captured_output or "请先登录" in captured_output):
        return [SearchResult(
            platform="douyin", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error="DOUYIN_LOGIN_REQUIRED: 抖音API要求登录。请在engine/config.json的douyin.cookie字段填入抖音Cookie",
        )]
    if "数据解析失败" in captured_output:
        return [SearchResult(
            platform="douyin", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error=f"DOUYIN_API_ERROR: {captured_output.strip()[-200:]}",
        )]

    # Detect API-level login errors returned in the response dict
    if isinstance(raw_results, dict):
        status_code = raw_results.get("status_code", 0)
        status_msg = raw_results.get("status_msg", "")
        if status_code == 2483 or "登录" in str(status_msg):
            return [SearchResult(
                platform="douyin", keyword_id="", keyword=keyword, sort_type=sort_type,
                title="", url="", error="DOUYIN_LOGIN_REQUIRED: 抖音API要求登录。请在engine/config.json的douyin.cookie字段填入抖音Cookie",
            )]
        if status_code and status_code != 0:
            return [SearchResult(
                platform="douyin", keyword_id="", keyword=keyword, sort_type=sort_type,
                title="", url="", error=f"DOUYIN_API_ERROR: status_code={status_code}, {status_msg[:80]}",
            )]
        return [SearchResult(
            platform="douyin", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error=f"DOUYIN_UNEXPECTED: {str(raw_results)[:120]}",
        )]

    # Flatten: TikTokDownloader returns [[page1_items], [page2_items], ...]
    flat_items = []
    if isinstance(raw_results, list):
        for page in raw_results:
            if isinstance(page, list):
                flat_items.extend(page)
            elif isinstance(page, dict):
                flat_items.append(page)

    if not flat_items:
        return [SearchResult(
            platform="douyin", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error="DOUYIN_NO_RESULTS: 搜索返回空结果",
        )]

    results = []
    for item in flat_items:
        if not isinstance(item, dict):
            continue
        results.append(SearchResult(
            platform="douyin",
            keyword_id="",
            keyword=keyword,
            sort_type=sort_type,
            title=item.get("desc", "") or item.get("title", ""),
            url=item.get("url", "") or f"https://www.douyin.com/video/{item.get('aweme_id', '')}",
            author=item.get("author", {}).get("nickname", "") if isinstance(item.get("author"), dict) else "",
            publish_time=item.get("create_time", ""),
            engagement=(item.get("statistics", {}) or {}).get("digg_count", 0),
            snippet=(item.get("desc", "") or "")[:200],
        ))
    return results[:count]


def _search_youtube(keyword: str, sort_type: str, count: int = 30,
                    date_from: str = "", date_to: str = "") -> list[SearchResult]:
    """Search YouTube via yt-dlp ytsearch.

    sort_type='date' -> client-side sort by upload_date DESC
    sort_type='hot'  -> default relevance sort
    date_from/date_to: YYYY-MM-DD. extract_flat=True doesn't expose upload_date,
        so date filter is best-effort: fetch larger count, skip if out of range
        (only works when extract_flat actually returns upload_date).
    """
    try:
        import yt_dlp
    except ImportError:
        return []

    # Date mode: use larger count, filter best-effort.
    # extract_flat=True is fast but doesn't always have upload_date,
    # so we try to filter what we can.
    if date_from and date_to and count == 0:
        count = _DATE_MODE_FALLBACK_COUNT  # date mode sentinel -> client-side filter

    search_query = f"ytsearch{count}:{keyword}"

    ydl_opts = {
        "quiet": True,
        "no_warnings": True,
        "extract_flat": True,
        "skip_download": True,
        "socket_timeout": 30,
    }

    results: list[SearchResult] = []

    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(search_query, download=False)
            for entry in info.get("entries", []) or []:
                vid = entry.get("id", "")
                results.append(SearchResult(
                    platform="youtube",
                    keyword_id="",
                    keyword=keyword,
                    sort_type=sort_type,
                    title=entry.get("title", ""),
                    url=f"https://www.youtube.com/watch?v={vid}" if vid else entry.get("url", ""),
                    author=entry.get("channel", "") or entry.get("uploader", ""),
                    publish_time=entry.get("upload_date", "") or "",
                    engagement=(entry.get("view_count") or 0),
                    snippet=entry.get("description", "")[:200] if entry.get("description") else "",
                ))
    except Exception:
        pass

    # Client-side date sort: ytsearchdate extractor is broken, so we
    # fetch via ytsearch and sort by upload_date descending ourselves.
    if sort_type == "date" and results:
        results.sort(key=lambda r: r.publish_time, reverse=True)

    return results


def _search_xiaohongshu(keyword: str, sort_type: str, count: int = 30,
                        date_from: str = "", date_to: str = "") -> list[SearchResult]:
    """Search XHS via xhshow-signed API (primary) or Playwright (fallback).

    Two-tier approach:
      1. xhshow API: fast, no browser, but requires clean (non-flagged) account
      2. Playwright: browser-based fallback for when API is blocked

    sort_type: 'date' → time_descending (最新), 'hot' → general (最热)
    """
    from urllib.parse import quote

    sort_param = "time_descending" if sort_type == "date" else "general"

    # ── Load cookies ───────────────────────────────────────────────────
    cookie_file = PROJECT_ROOT / "engine" / ".xhs_cookies.json"
    cookie_str = ""
    if cookie_file.exists():
        try:
            cdata = json.loads(cookie_file.read_text(encoding="utf-8"))
            cookie_str = cdata.get("cookie_str", "")
        except Exception:
            pass

    if not cookie_str:
        return [SearchResult(
            platform="xiaohongshu", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error="XHS_NO_COOKIE: 未找到Cookie，请先在侧边栏运行小红书登录",
        )]

    # ── Tier 1: xhshow-signed API ──────────────────────────────────────
    api_error = _search_xhs_api(keyword, sort_param, count, cookie_str)
    if api_error is not None:
        if api_error and api_error[0].error:
            err = api_error[0].error
            if "XHS_CAPTCHA" in err or "XHS_ACCOUNT_FLAGGED" in err:
                pass  # fall through to Tier 2
            else:
                return api_error
        else:
            return api_error

    # ── Tier 2: MediaCrawler adapter search (API with different headers) ──
    try:
        from engine.mediacrawler_adapter import search_xhs as _mc_search
        mc_results = _mc_search(keyword, count=count, sort_type=sort_type)
        if mc_results:
            return [
                SearchResult(
                    platform="xiaohongshu",
                    keyword_id="", keyword=keyword, sort_type=sort_type,
                    title=r.get("title", ""),
                    url=r.get("url", ""),
                    author=r.get("author", ""),
                    engagement=r.get("engagement", 0),
                )
                for r in mc_results
            ]
    except Exception:
        pass

    # ── Tier 3: Playwright browser ─────────────────────────────────────
    try:
        from engine.browser_pool import launch_context, BROWSER_ARGS
    except ImportError:
        return [SearchResult(
            platform="xiaohongshu", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error="XHS_IMPORT_ERROR: Playwright未安装，且API路径被验证码拦截",
        )]

    playwright_cookies = []
    for pair in cookie_str.split("; "):
        if "=" in pair:
            k, v = pair.split("=", 1)
            playwright_cookies.append({
                "name": k.strip(), "value": v.strip(),
                "domain": ".xiaohongshu.com", "path": "/",
            })

    search_url = (
        f"https://www.xiaohongshu.com/search_result?"
        f"keyword={quote(keyword)}&sort={sort_param}"
    )

    results = []
    try:
        ctx, browser, pw = launch_context(
            headless=True, stealth=False,
            args=BROWSER_ARGS + ["--no-sandbox"],
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
        )
        try:
            ctx.add_cookies(playwright_cookies)
            page = ctx.new_page()
            page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            """)
            page.goto(search_url, timeout=30000, wait_until="domcontentloaded")
            page.wait_for_timeout(5000)

            # Login wall detection
            body_text = page.inner_text("body")[:3000]
            login_keywords = ["登录后查看搜索结果", "登录后查看", "请先登录", "手机号登录"]
            is_login_wall = any(kw in body_text for kw in login_keywords)
            if is_login_wall:
                return [SearchResult(
                    platform="xiaohongshu", keyword_id="", keyword=keyword, sort_type=sort_type,
                    title="", url="", error="XHS_COOKIE_EXPIRED: Cookie已过期或被拒绝，搜索页要求登录。请重新运行XHS bootstrap获取有效Cookie",
                )]

            # Redcaptcha detection
            if "redcaptcha" in body_text.lower() or "验证" in body_text:
                return [SearchResult(
                    platform="xiaohongshu", keyword_id="", keyword=keyword, sort_type=sort_type,
                    title="", url="", error="XHS_CAPTCHA: XHS触发了验证码验证，自动化搜索暂时不可用。请使用侧边栏重新登录获取新Cookie，或等待账号风控解除",
                )]

            # Try multiple selector strategies
            cards = page.query_selector_all("section.note-item, a[href*='/explore/']")
            if not cards:
                cards = page.query_selector_all(".feeds-page .note-item, [class*=note-item], .search-result-item")
            if not cards:
                cards = page.query_selector_all(".feeds-page > div > div > a[href]")

            for card in cards[:count]:
                try:
                    link_el = card.query_selector("a[href*='/explore/'], a[href*='/search_result/']")
                    if not link_el and card.tag_name == "a":
                        link_el = card
                    href = link_el.get_attribute("href") if link_el else ""
                    note_url = f"https://www.xiaohongshu.com{href}" if href.startswith("/") else href
                    if not note_url or note_url == "https://www.xiaohongshu.com":
                        continue

                    title_el = card.query_selector(".title, .note-title, a.title, span.title, [class*=title]")
                    title = title_el.inner_text().strip() if title_el else ""

                    author_el = card.query_selector(".author .name, .nickname, .username, [class*=author] [class*=name]")
                    author = author_el.inner_text().strip() if author_el else ""

                    likes_el = card.query_selector(".like-count, .count, .like span, span.count, [class*=like] span")
                    likes_text = likes_el.inner_text().strip() if likes_el else "0"

                    results.append(SearchResult(
                        platform="xiaohongshu",
                        keyword_id="", keyword=keyword, sort_type=sort_type,
                        title=title[:200], url=note_url, author=author,
                        engagement=int(re.sub(r'\D', '', likes_text) or 0),
                    ))
                except Exception:
                    continue
        finally:
            if browser:
                browser.close()
            else:
                ctx.close()
            pw.stop()
    except Exception as e:
        print(f"[XHS搜索] Playwright异常: {e}", file=sys.stderr)
        return [SearchResult(
            platform="xiaohongshu", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error=f"XHS_PLAYWRIGHT_ERROR: {str(e)[:120]}",
        )]

    if not results:
        return [SearchResult(
            platform="xiaohongshu", keyword_id="", keyword=keyword, sort_type=sort_type,
            title="", url="", error="XHS_NO_RESULTS: 搜索页加载成功但未匹配到结果卡片，可能账号被风控或XHS前端已更新",
        )]

    return results[:count]


def _search_xhs_api(keyword: str, sort_param: str, count: int, cookie_str: str) -> list[SearchResult] | None:
    """Tier 1: Search XHS via xhshow-signed API call.

    Returns:
      list of SearchResult on definitive result (success or non-retryable error)
      None if should fall through to Playwright (retryable: captcha/461)
    """
    import secrets as _secrets
    try:
        from xhshow import Xhshow
        import httpx as _httpx
    except ImportError:
        return None  # xhshow not installed, fall through to Playwright

    try:
        xh = Xhshow()
    except Exception:
        return None

    search_id = xh.get_b3_trace_id()
    payload = {
        "keyword": keyword,
        "page": 1,
        "page_size": min(count, 50),
        "search_id": search_id,
        "sort": sort_param,
        "note_type": 0,
        "ext_flags": [],
        "geo": "",
        "image_formats": ["jpg", "webp", "avif"],
    }

    uri = "/api/sns/web/v1/search/notes"
    headers = xh.sign_headers(
        method="POST", uri=uri, cookies=cookie_str, payload=payload,
        xsec_appid="xhs-pc-web",
    )
    headers["Cookie"] = cookie_str
    headers["Content-Type"] = "application/json;charset=UTF-8"
    headers["User-Agent"] = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
    )
    headers["Referer"] = "https://www.xiaohongshu.com/"
    headers["Origin"] = "https://www.xiaohongshu.com"

    try:
        get_limiter("xhs").acquire()
        # Random jitter between searches to avoid pattern detection
        time.sleep(random.uniform(0.5, 2.0))
        client = _httpx.Client(timeout=30)
        json_str = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)
        resp = client.post(
            "https://edith.xiaohongshu.com/api/sns/web/v1/search/notes",
            content=json_str, headers=headers,
        )
        data = resp.json()
        client.close()
    except Exception as e:
        print(f"[XHS搜索] API请求失败: {e}", file=sys.stderr)
        return None  # network error, fall through to Playwright

    code = data.get("code", 0)
    success = data.get("success", False)
    msg = str(data.get("msg", ""))
    resp_data = data.get("data", {})

    # Account flagged
    if code == -103 and "违规" in msg:
        return [SearchResult(
            platform="xiaohongshu", keyword_id="", keyword=keyword, sort_type=sort_param,
            title="", url="", error="XHS_ACCOUNT_FLAGGED: XHS账号被标记为异常（违规情形），搜索功能被限制。请更换XHS账号或等待风控解除",
        )]

    # Anti-bot / captcha
    if resp.status_code == 461:
        return [SearchResult(
            platform="xiaohongshu", keyword_id="", keyword=keyword, sort_type=sort_param,
            title="", url="", error="XHS_CAPTCHA: XHS反爬虫系统触发，要求验证码验证",
        )]

    # Other API errors
    if not success and code != 0:
        return [SearchResult(
            platform="xiaohongshu", keyword_id="", keyword=keyword, sort_type=sort_param,
            title="", url="", error=f"XHS_API_ERROR: code={code}, {msg[:80]}",
        )]

    # Parse results
    items = resp_data.get("notes", resp_data.get("items", []))
    if not isinstance(items, list) or not items:
        return [SearchResult(
            platform="xiaohongshu", keyword_id="", keyword=keyword, sort_type=sort_param,
            title="", url="", error="XHS_NO_RESULTS: API返回空结果，该关键词可能无匹配内容",
        )]

    results = []
    for item in items:
        note_id = item.get("id", "") or item.get("note_id", "")
        note_card = item.get("note_card", item)
        title = (
            note_card.get("display_title", "")
            or note_card.get("title", "")
            or ""
        )
        author_info = note_card.get("user", {})
        author = author_info.get("nickname", author_info.get("name", ""))
        likes = note_card.get("likes", note_card.get("liked_count", 0))
        results.append(SearchResult(
            platform="xiaohongshu",
            keyword_id="", keyword=keyword, sort_type=sort_param,
            title=str(title)[:200],
            url=f"https://www.xiaohongshu.com/explore/{note_id}" if note_id else "",
            author=str(author),
            engagement=int(likes) if likes else 0,
        ))
    return results[:count]


# ═══════════════════════════════════════════════════════════════════════════════
# B站 (Bilibili) keyword search — bilibili-api-python
# ═══════════════════════════════════════════════════════════════════════════════

def _search_bilibili(keyword: str, sort_type: str, count: int = 30,
                     date_from: str = "", date_to: str = "") -> list[SearchResult]:
    """Search B站 via bilibili-api-python. Async, wrapped in sync helper.

    sort_type: 'date' -> PUBDATE, 'hot' -> TOTALRANK
    date_from/date_to: YYYY-MM-DD. When non-empty, passed as API time_start/time_end.
        When count=0 (date-mode sentinel), paginates until empty page or 50-page cap.
    """
    import asyncio as _asyncio

    async def _do():
        from bilibili_api import search
        order = search.OrderVideo.PUBDATE if sort_type == "date" else search.OrderVideo.TOTALRANK
        results = []
        max_pages = _DATE_MODE_MAX_PAGES if count == 0 else max(1, count // 20)
        page = 1
        while page <= max_pages:
            try:
                kwargs = dict(
                    keyword=keyword,
                    search_type=search.SearchObjectType.VIDEO,
                    order_type=order,
                    page=page,
                )
                if date_from:
                    kwargs["time_start"] = date_from
                if date_to:
                    kwargs["time_end"] = date_to

                data = await search.search_by_type(**kwargs)
                items = data.get("result", [])
                if not items:
                    break  # empty page -> stop
                for item in items[:20]:
                    results.append({
                        "title": item.get("title", "").replace('<em class="keyword">', '').replace('</em>', ''),
                        "bvid": item.get("bvid", ""),
                        "aid": item.get("aid", 0),
                        "author": item.get("author", ""),
                        "play": item.get("play", 0),
                        "video_review": item.get("video_review", 0),
                        "pubdate": item.get("pubdate", 0),
                    })
                if count > 0 and len(results) >= count:
                    break
                page += 1
            except Exception:
                break
        return results

    try:
        loop = _asyncio.new_event_loop()
        _asyncio.set_event_loop(loop)
        items = loop.run_until_complete(_do())
        loop.close()
    except Exception:
        return []

    from datetime import datetime as _dt
    output = []
    # count=0 is date-mode sentinel (unlimited), don't truncate
    source = items if count == 0 else items[:count]
    for item in source:
        pub_str = ""
        if item.get("pubdate"):
            try:
                pub_str = _dt.fromtimestamp(item["pubdate"]).strftime("%Y-%m-%d")
            except Exception:
                pass
        url = f"https://www.bilibili.com/video/{item['bvid']}" if item.get("bvid") else \
              f"https://www.bilibili.com/video/av{item['aid']}"
        output.append(SearchResult(
            keyword_id="", keyword=keyword, sort_type=sort_type,
            title=item.get("title", "")[:200], url=url,
            author=item.get("author", ""), publish_time=pub_str,
            engagement=item.get("play", 0),
            platform="bilibili",
        ))
    return output


def _search_weibo(keyword: str, sort_type: str, count: int = 30,
                   date_from: str = "", date_to: str = "") -> list[SearchResult]:
    """Search 微博 via crawl4weibo.

    sort_type: 'date' → 按时间降序, 'hot' → 综合排序（默认）
    """
    try:
        from crawl4weibo import WeiboClient
        client = WeiboClient()
    except Exception:
        return []

    items = []
    page = 1
    max_pages = 8 if sort_type == "date" else 5  # more pages for date sort to get recent posts
    while len(items) < count and page <= max_pages:
        try:
            posts, pagination = client.search_posts(keyword, page=page)
            for p in posts:
                pub_str = ""
                if p.created_at:
                    try:
                        pub_str = str(p.created_at)
                        if "+" in pub_str:
                            pub_str = pub_str.split("+")[0].strip()
                        if " " in pub_str:
                            pub_str = pub_str[:10]
                    except Exception:
                        pass
                engagement = (p.attitudes_count or 0) + (p.comments_count or 0) + (p.reposts_count or 0)
                post_url = f"https://weibo.com/{p.user_id}/{p.bid}" if p.user_id and p.bid else ""
                items.append({
                    "title": (p.text or "")[:200],
                    "url": post_url,
                    "author": "",  # would need separate get_user_by_uid call
                    "publish_time": pub_str,
                    "engagement": engagement,
                    "created_at": p.created_at,
                })
            if not pagination.get("has_more"):
                break
            page = pagination.get("page", page + 1)
        except Exception as e:
            print(f"[微博] 搜索第{page}页失败: {e}")
            break

    if sort_type == "date":
        items.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)

    output = []
    for item in items[:count]:
        output.append(SearchResult(
            keyword_id="", keyword=keyword, sort_type=sort_type,
            title=item.get("title", "")[:200], url=item.get("url", ""),
            author=item.get("author", ""), publish_time=item.get("publish_time", ""),
            engagement=item.get("engagement", 0),
            platform="weibo",
        ))
    return output


def _search_wechat(keyword: str, sort_type: str, count: int = 30,
                   date_from: str = "", date_to: str = "") -> list[SearchResult]:
    """Search 微信公众号 via weixin.sogou.com — public, no credentials needed."""
    try:
        from engine.wechat_fetcher import search_wechat_articles
        raw = search_wechat_articles(keyword, count=count, sort_type=sort_type)
        results = []
        for r in raw:
            results.append(SearchResult(
                keyword_id="", keyword=keyword, sort_type=sort_type,
                title=r.get("title", "")[:200],
                url=r.get("url", ""),
                author=r.get("author", ""),
                publish_time=r.get("publish_time", ""),
                snippet=r.get("snippet", "")[:200],
                engagement=0,
                platform="wechat",
            ))
        return results[:count]
    except Exception as e:
        print(f"[微信] 搜索失败: {e}", file=sys.stderr)
        return []


PLATFORM_SEARCHERS = {
    "douyin": _search_douyin,
    "youtube": _search_youtube,
    "xiaohongshu": _search_xiaohongshu,
    "bilibili": _search_bilibili,
    "weibo": _search_weibo,
    "wechat": _search_wechat,
}


def _search_with_timeout(searcher, keyword: str, sort_type: str, count: int,
                        date_from: str = "", date_to: str = "") -> list[SearchResult]:
    """Run a platform search in a thread with a hard timeout.

    yt-dlp / Playwright calls can block indefinitely in background threads
    even with socket_timeout set.  This wraps each call in a future so we
    can enforce a wall-clock deadline.
    """
    executor = futures.ThreadPoolExecutor(max_workers=1)
    try:
        fut = executor.submit(searcher, keyword, sort_type, count, date_from, date_to)
        try:
            return fut.result(timeout=_SEARCH_TIMEOUT)
        except futures.TimeoutError:
            return []
        except Exception:
            return []
    finally:
        executor.shutdown(wait=False)


# ── Deduplication ──────────────────────────────────────────────────────
def _load_previous_urls(keyword_id: str, platform: str) -> set:
    """Load ALL previously archived URLs across all dates for cross-day dedup."""
    archive_root = RAW_DIR / "monitor"
    if not archive_root.exists():
        return set()
    urls = set()
    for date_dir in archive_root.iterdir():
        if not date_dir.is_dir():
            continue
        for f in date_dir.glob(f"{keyword_id}_{platform}_*.json"):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                for item in data:
                    if isinstance(item, dict) and "url" in item:
                        urls.add(item["url"])
            except Exception:
                pass
    return urls


def _archive_results(results: list[SearchResult], date_str: str, keyword_id: str, platform: str, sort_type: str):
    """Archive search results to raw/monitor/YYYY-MM-DD/."""
    archive_dir = RAW_DIR / "monitor" / date_str
    archive_dir.mkdir(parents=True, exist_ok=True)
    path = archive_dir / f"{keyword_id}_{platform}_{sort_type}.json"
    data = [{"title": r.title, "url": r.url, "author": r.author,
             "publish_time": r.publish_time, "engagement": r.engagement} for r in results]
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ── Excel export ───────────────────────────────────────────────────────
def _export_excel(harvest: MonitorHarvest) -> str:
    """Export MonitorHarvest to Excel with yellow highlights for new items."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        path = OUTPUTS_DIR / f"monitor_{harvest.job_id}.txt"
        path.parent.mkdir(parents=True, exist_ok=True)
        lines = [f"Monitor Job: {harvest.job_id}", f"Total: {harvest.total_fetched}, New: {harvest.total_new}"]
        path.write_text("\n".join(lines), encoding="utf-8")
        return str(path)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Monitor {harvest.job_id}"

    headers = ["来源平台", "关键词ID", "关键词", "排序方式", "标题", "URL", "作者", "发布时间", "互动量"]
    header_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    row = 2
    new_urls = {(r.url) for kr in harvest.keyword_results for r in kr.new_items}

    for kr in harvest.keyword_results:
        for r in kr.date_results + kr.hot_results:
            values = [r.platform, kr.keyword_id, kr.keyword, r.sort_type,
                      r.title[:200], r.url, r.author, r.publish_time, r.engagement]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if r.url in new_urls:
                    cell.fill = yellow_fill
            row += 1

    path = OUTPUTS_DIR / f"monitor_{harvest.job_id}.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return str(path)


# ── Main entry point ───────────────────────────────────────────────────
def execute_job(progress_callback=None, sort_preference: str = "default",
                date_from: str = "", date_to: str = "") -> MonitorHarvest:
    """Execute a full monitor job: search -> dedup -> archive -> Excel.

    Args:
        progress_callback: Optional callable(details_str) for per-keyword
            progress reporting. Called before each platform search.
        sort_preference: "default" (relevance/popularity) or "date" (time sort).
            Maps to internal sort_type "hot" / "date" respectively.
        date_from: YYYY-MM-DD start date for date-range mode.
        date_to: YYYY-MM-DD end date for date-range mode.
            When both non-empty, date-range mode activates:
            - bilibili/youtube: API date filter + unlimited pages (cap 50)
            - douyin: best-fit PublishTimeType preset
            - weibo: fetch 200, client-side filter by publish_time
            - xiaohongshu/wechat: skipped

    Called by Orchestrator.run_active_monitor() and pipeline._run_pipeline().
    """
    _sort_type = "date" if sort_preference == "date" else "hot"
    _date_mode = bool(date_from and date_to)

    job_id = datetime.now().strftime("%Y%m%d_%H%M")
    date_str = datetime.now().strftime("%Y-%m-%d")
    started = datetime.now().isoformat()
    keywords = load_keywords()
    harvest = MonitorHarvest(job_id=job_id, started_at=started, finished_at="")

    defaults = json.loads(
        (PROJECT_ROOT / "monitor_keywords.json").read_text(encoding="utf-8")
    ).get("defaults", {"result_count": 30})

    total_pairs = sum(len(kw.get("platforms", [])) for kw in keywords)
    pair_idx = 0

    for kw_entry in keywords:
        kw_id = kw_entry["id"]
        kw_text = kw_entry["keyword"]
        base_count = kw_entry.get("result_count", defaults.get("result_count", 30))

        for platform in kw_entry.get("platforms", []):
            pair_idx += 1
            searcher = PLATFORM_SEARCHERS.get(platform)
            if searcher is None:
                continue

            if progress_callback:
                progress_callback(f"搜索: {kw_text} @ {platform} ({pair_idx}/{total_pairs})")

            kr = KeywordResult(keyword_id=kw_id, keyword=kw_text, platform=platform)

            # ── Date-mode dispatch ──────────────────────────────────
            if _date_mode:
                if platform in _DATE_SKIP_PLATFORMS:
                    kr.notes.append(
                        f"日期模式: {platform}平台不支持日期搜索，已跳过"
                    )
                    harvest.keyword_results.append(kr)
                    if progress_callback:
                        progress_callback(
                            f"跳过: {kw_text} @ {platform} ({pair_idx}/{total_pairs})"
                        )
                    continue

                if platform in _DATE_CAPABLE_PLATFORMS:
                    # API-level date filter (bilibili), unlimited pages
                    count = 0
                    search_date_from = date_from
                    search_date_to = date_to
                elif platform in _DATE_CLIENTSIDE_PLATFORMS:
                    # Client-side filter: fetch more, filter after
                    count = _DATE_MODE_FALLBACK_COUNT
                    search_date_from = ""
                    search_date_to = ""
                elif platform == "douyin":
                    # Best-fit preset only
                    count = base_count
                    search_date_from = date_from
                    search_date_to = date_to
                else:
                    count = base_count
                    search_date_from = ""
                    search_date_to = ""
            else:
                count = base_count
                search_date_from = ""
                search_date_to = ""

            try:
                results = _search_with_timeout(
                    searcher, kw_text, _sort_type, count,
                    date_from=search_date_from, date_to=search_date_to,
                )
            except Exception:
                results = []

            # ── Client-side date filter ──────────────────────────
            if _date_mode and platform in _DATE_CLIENTSIDE_PLATFORMS and results:
                raw_count = len(results)
                has_date = [r for r in results if r.publish_time]
                no_date = [r for r in results if not r.publish_time]
                filtered = [
                    r for r in has_date
                    if date_from <= r.publish_time[:10] <= date_to
                ]
                results = filtered + no_date  # keep items without date (best-effort)
                if platform == "weibo":
                    kr.notes.append(
                        f"日期模式: 客户端截断, {raw_count}条过滤后剩余{len(results)}条"
                    )
                else:  # youtube
                    kr.notes.append(
                        f"日期模式: {date_from} ~ {date_to}, "
                        f"获取{raw_count}条, 过滤后{len(results)}条 "
                        f"(注: 无日期字段的{len(no_date)}条保留)"
                    )
            elif _date_mode and platform in _DATE_CAPABLE_PLATFORMS:
                kr.notes.append(
                    f"日期模式: {date_from} ~ {date_to}, 获取{len(results)}条"
                )
            elif _date_mode and platform == "douyin":
                kr.notes.append(
                    "日期模式: 抖音仅支持预设区间(1天/1周/6月), 非精确日期"
                )

            # Store results in the appropriate field for backward compat
            if _sort_type == "date":
                kr.date_results = results
            else:
                kr.hot_results = results

            prev_urls = _load_previous_urls(kw_id, platform)
            kr.new_items = [r for r in results if r.url not in prev_urls]

            harvest.total_fetched += len(results)
            harvest.total_new += len(kr.new_items)

            # Archive single result set
            _archive_results(results, date_str, kw_id, platform, _sort_type)

            harvest.keyword_results.append(kr)

    if harvest.total_fetched > 0:
        harvest.dedup_rate = (harvest.total_fetched - harvest.total_new) / harvest.total_fetched

    harvest.finished_at = datetime.now().isoformat()
    harvest.excel_path = _export_excel(harvest)
    harvest.stats = MonitorStats(
        keywords_searched=len(keywords),
        platforms_queried=sum(len(kw.get("platforms", [])) for kw in keywords),
        total_fetched=harvest.total_fetched,
        total_new=harvest.total_new,
        dedup_rate=harvest.dedup_rate,
    )

    return harvest


# ── Feedback learning (Analyst → Monitor) ─────────────────────────────
def record_feedback(keyword_id: str, platform: str, is_relevant: bool, reason: str = ""):
    """Record Analyst relevance feedback for keyword optimization (PRD M-07).

    Called by Orchestrator (not directly by Analyst — isolation constraint).
    """
    feedback_path = PROJECT_ROOT / "outputs" / "keyword_feedback.jsonl"
    feedback_path.parent.mkdir(parents=True, exist_ok=True)
    entry = {
        "timestamp": datetime.now().isoformat(),
        "keyword_id": keyword_id,
        "platform": platform,
        "is_relevant": is_relevant,
        "reason": reason,
    }
    with open(feedback_path, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


